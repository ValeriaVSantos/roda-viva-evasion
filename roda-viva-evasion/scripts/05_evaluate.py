"""
avaliar_metodos.py
==================
Combina os resultados dos 3 métodos e gera tabela de avaliação
com precision / recall / F1 por classe e macro, mais Cohen's Kappa.

Uso (após rodar os 3 métodos):
    python avaliar_metodos.py \
        --sim  resultados_similaridade.csv \
        --nli  resultados_nli.csv \
        --llm  resultados_llm.csv \
        --output tabela_resultados.xlsx
"""

import argparse, csv
from collections import defaultdict

CLASSES = ["RESPONDE", "PARCIAL", "ESQUIVA"]

def ler_predicoes(caminho, coluna_pred):
    gold, pred = [], []
    with open(caminho, encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            g = row.get("GOLD","").strip().upper()
            p = row.get(coluna_pred,"").strip().upper()
            if g in CLASSES and p in CLASSES:
                gold.append(g)
                pred.append(p)
    return gold, pred

def metricas_por_classe(gold, pred):
    from sklearn.metrics import precision_recall_fscore_support, cohen_kappa_score
    p, r, f, _ = precision_recall_fscore_support(gold, pred, labels=CLASSES, zero_division=0)
    pm, rm, fm, _ = precision_recall_fscore_support(gold, pred, average="macro", zero_division=0)
    kappa = cohen_kappa_score(gold, pred)
    acuracia = sum(g==p for g,p in zip(gold,pred)) / len(gold)
    return {
        "por_classe": {cls: {"P": p[i], "R": r[i], "F1": f[i]}
                       for i, cls in enumerate(CLASSES)},
        "macro":  {"P": pm, "R": rm, "F1": fm},
        "kappa":  kappa,
        "acc":    acuracia,
        "n":      len(gold),
    }

def imprimir_tabela(nome, metricas):
    print(f"\n{'='*60}")
    print(f"  {nome}  (n={metricas['n']})")
    print(f"{'='*60}")
    print(f"{'Classe':<12} {'Prec':>6} {'Rec':>6} {'F1':>6}")
    print(f"{'-'*34}")
    for cls in CLASSES:
        m = metricas["por_classe"][cls]
        print(f"{cls:<12} {m['P']:>6.3f} {m['R']:>6.3f} {m['F1']:>6.3f}")
    print(f"{'-'*34}")
    m = metricas["macro"]
    print(f"{'MACRO':<12} {m['P']:>6.3f} {m['R']:>6.3f} {m['F1']:>6.3f}")
    print(f"  Acurácia:      {metricas['acc']:.3f}")
    print(f"  Cohen's Kappa: {metricas['kappa']:.3f}")

def salvar_xlsx(resultados, caminho):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    wb = openpyxl.Workbook()

    # ── Aba 1: Resumo geral ───────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumo"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    cabecalho = ["Método","Classe","Precision","Recall","F1-score","Kappa","Acurácia"]
    for col, h in enumerate(cabecalho, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center")

    cores_classe = {"RESPONDE": "C6EFCE", "PARCIAL": "FFEB9C", "ESQUIVA": "FFC7CE"}
    row_idx = 2
    for nome, metricas in resultados.items():
        for cls in CLASSES:
            m = metricas["por_classe"][cls]
            ws.cell(row=row_idx, column=1, value=nome)
            ws.cell(row=row_idx, column=2, value=cls).fill = PatternFill("solid", fgColor=cores_classe[cls])
            ws.cell(row=row_idx, column=3, value=round(m["P"], 3))
            ws.cell(row=row_idx, column=4, value=round(m["R"], 3))
            ws.cell(row=row_idx, column=5, value=round(m["F1"], 3))
            ws.cell(row=row_idx, column=6, value=round(metricas["kappa"], 3))
            ws.cell(row=row_idx, column=7, value=round(metricas["acc"], 3))
            row_idx += 1
        # Linha macro
        m = metricas["macro"]
        ws.cell(row=row_idx, column=1, value=nome)
        ws.cell(row=row_idx, column=2, value="MACRO").font = Font(bold=True)
        ws.cell(row=row_idx, column=3, value=round(m["P"], 3))
        ws.cell(row=row_idx, column=4, value=round(m["R"], 3))
        ws.cell(row=row_idx, column=5, value=round(m["F1"], 3))
        ws.cell(row=row_idx, column=6, value=round(metricas["kappa"], 3))
        ws.cell(row=row_idx, column=7, value=round(metricas["acc"], 3))
        row_idx += 2  # linha em branco entre métodos

    for col in ["A","B","C","D","E","F","G"]:
        ws.column_dimensions[col].width = 16
    ws.freeze_panes = "A2"

    wb.save(caminho)
    print(f"\n[OK] Tabela salva: {caminho}")

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--sim",    default="resultados_similaridade.csv")
    parser.add_argument("--nli",    default="resultados_nli.csv")
    parser.add_argument("--llm",    default="resultados_llm.csv")
    parser.add_argument("--output", default="tabela_resultados.xlsx")
    args = parser.parse_args()

    resultados = {}

    # Pré-anotação como baseline
    g0, p0 = ler_predicoes(args.sim, "GOLD")  # gold vs gold = perfeito, usar pre_anotacao
    # Lemos a coluna PRE_ANOTACAO do arquivo de similaridade como baseline
    gold_base, pre_base = [], []
    with open(args.sim, encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            g = row.get("GOLD","").strip().upper()
            p = row.get("PRE_ANOTACAO","").strip().upper()
            if g in CLASSES and p in CLASSES:
                gold_base.append(g)
                pre_base.append(p)
    if gold_base:
        resultados["Baseline (heurística)"] = metricas_por_classe(gold_base, pre_base)

    # Método 1: similaridade
    try:
        gold, pred = ler_predicoes(args.sim, "METODO1_SIM")
        if gold:
            resultados["Método 1: Similaridade"] = metricas_por_classe(gold, pred)
    except FileNotFoundError:
        print(f"[AVISO] {args.sim} não encontrado — pulando Método 1")

    # Método 2: NLI
    try:
        gold, pred = ler_predicoes(args.nli, "METODO2_NLI")
        if gold:
            resultados["Método 2: NLI"] = metricas_por_classe(gold, pred)
    except FileNotFoundError:
        print(f"[AVISO] {args.nli} não encontrado — pulando Método 2")

    # Método 3: LLM
    try:
        gold, pred = ler_predicoes(args.llm, "METODO3_LLM")
        if gold:
            resultados["Método 3: LLM zero-shot"] = metricas_por_classe(gold, pred)
    except FileNotFoundError:
        print(f"[AVISO] {args.llm} não encontrado — pulando Método 3")

    if not resultados:
        print("[ERRO] Nenhum arquivo de resultado encontrado.")
        return

    for nome, metricas in resultados.items():
        imprimir_tabela(nome, metricas)

    salvar_xlsx(resultados, args.output)

if __name__ == "__main__":
    main()
